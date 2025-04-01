# coding: utf_8
# 纯文本文件 student.txt为学生信息，请将上述内容写到 student.xls 文件中

import json, os
from openpyxl import Workbook

def txt2xlsx(txtFile, xlsxFile = None):
	y = -1 # 成功插入数据的行数
	if os.path.exists(txtFile):
		y = 0
	else:
		return y
	with open(txtFile, "rt", encoding="utf_8") as txt:
		data = json.load(txt)
		print (data)
		xlsx = Workbook()
		sheet1 = xlsx.active
		sheet1.title = os.path.splitext(txtFile)[0]
		row = 1
		col = 1
		for i in range(1, len(data)+1):
			sheet1.cell(row = i, column = col).value = i
			L = data[str(i)]
			for j in range(0, len(L)):
				sheet1.cell(row = i, column = j + 2).value = L[j]
			y += 1
		if xlsxFile == None:
			xlsxFile = sheet1.title + ".xlsx"
		xlsx.save(xlsxFile)
	return y

if __name__ == "__main__":
	y = txt2xlsx("student.txt")
	print(y, "行数据")
